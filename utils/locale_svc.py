import json
from dataclasses import dataclass
from pathlib import Path
from collections import namedtuple
from functools import cached_property


import openpyxl
from openpyxl import load_workbook
from chardet.universaldetector import UniversalDetector



# JsonLocalizationDataRouter = namedtuple("JsonLocalizationDataRouter", ('json', 'json_dir_path', 'json_path', 'xlsx_path'))


@dataclass
class JsonLocalizationDataRouter:
    json_dir_path: Path
    json_path: Path
    xlsx_path: Path

    def __post_init__(self):
        if not self.xlsx_path.exists():
            self.xlsx_path.touch()
    
    @cached_property
    def flatten_json(self):
        return flatten_dict(json.loads(self.json_path.read_text()))
    


def get_decode_filetype(file_path):
    detector = UniversalDetector()
    with open(file_path, 'rb') as fh:
        for line in fh:
            detector.feed(line)
            if detector.done:
                break
        detector.close()
    return detector.result


def get_json_files_and_path(all_lang_path, lang_dir_name=None) -> [JsonLocalizationDataRouter]:
    result = []
    for lang_dir in all_lang_path.iterdir():
        if lang_dir_name is None or lang_dir.name.lower() == lang_dir_name.lower():
            dir_result = []
            for json_file in lang_dir.iterdir():
                if json_file.suffix == ".json" and "_new" not in json_file.stem:
                    json_path = json_file.resolve()
                    json_dir = (json_path / "..").resolve()
                    xlsx_path = json_dir / f"{json_path.stem}.xlsx"

                    dir_result.append(JsonLocalizationDataRouter(json_dir, json_path, xlsx_path))
            result.append(dir_result)

    return result


def flatten_dict(dictionary, parent_key='', sep=':'):
    flattened_dict = {}
    for key, value in dictionary.items():
        if isinstance(key, str) and key.isdigit():
            key = f"'{key}'"
        new_key = f"{parent_key}{sep}{key}" if parent_key else key
        if isinstance(value, dict):
            flattened_dict.update(flatten_dict(value, new_key, sep))
        elif isinstance(value, list):
            for i, item in enumerate(value):
                if isinstance(item, dict):

                    flattened_dict.update(flatten_dict(item, f"{new_key}{sep}{i}", sep))
                else:
                    flattened_dict[f"{new_key}{sep}{i}"] = item
        else:
            flattened_dict[new_key] = value
    return flattened_dict


def transform_dict_to_dict_with_list(d):
    if isinstance(d, dict):
        result = {}
        for k, v in d.items():
            if isinstance(v, dict):
                if all([i.isdigit() for i in v.keys()]):
                    result[k] = [transform_dict_to_dict_with_list(v) for k, v in sorted(v.items(), key=lambda x: int(x[0]))]
                else:
                    result[k] = transform_dict_to_dict_with_list(v)
            else:
                result[k] = v
        return result
    else:
        return d


def escape_dotc_for_digit_keys(d):
    if isinstance(d, dict):
        result = {}
        for k, v in d.items():
            if isinstance(v, dict):
                if all([i.isdigit() for i in v.keys()]):
                    result[k] = [escape_dotc_for_digit_keys(v) for k, v in sorted(v.items(), key=lambda x: int(x[0]))]
                else:
                    if "'" in k:
                        k = k.replace("'", "")
                    result[k] = escape_dotc_for_digit_keys(v)
            else:
                if "'" in k:
                    k = k.replace("'", "")
                result[k] = v
        return result
    else:
        return d


def unflatten_dict(data):
    result = {}

    for key, value in data.items():
        if key is None and value is None:
            continue
        if key is None:
            print(f"Skipped {key}:{value}")
            continue
        key = key.replace('.', ':')
        keys = key.split(':')
        current_dict = result

        for k in keys[:-1]:
            if k not in current_dict:
                current_dict[k] = {}

            current_dict = current_dict[k]

        last_key = keys[-1]

        if last_key not in current_dict:
            current_dict[last_key] = value
        else:
            current_value = current_dict[last_key]

            if isinstance(current_value, list):
                current_value.append(value)
            else:
                current_dict[last_key] = [current_value, value]

    return escape_dotc_for_digit_keys(transform_dict_to_dict_with_list(result))


def prepare_default_xlsx_file(lang_path: Path):
    _tmp = {i.json_path.stem: i for i in get_json_files_and_path(lang_path, "en")[0]}
    eng_path_lang_data = _tmp['en']
    eng_path_no_filename_data = _tmp.get('no_filename')


    all_langs_data: [JsonLocalizationDataRouter] = get_json_files_and_path(lang_path)

    for lang_metas in all_langs_data:
        for lang_meta in lang_metas:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            all_json_keys = list(lang_meta.flatten_json.keys())

            for row_index in range(len(all_json_keys)):
                key = all_json_keys[row_index]
                column1_cell = 'A{}'.format(row_index + 1)
                column2_cell = 'B{}'.format(row_index + 1)
                column3_cell = 'C{}'.format(row_index + 1)
                sheet[column1_cell] = key
                try:
                    sheet[column2_cell] = (eng_path_no_filename_data if "no_filename" in lang_meta.json_path.stem else eng_path_lang_data).flatten_json.get(key, "")
                except Exception as e:
                    print('qweqwe123213')
                sheet[column3_cell] = lang_meta.flatten_json[key] or ""

            workbook.save(lang_meta.xlsx_path)
            workbook.close()


def escape_slash(value):
    return value.replace('/', '\\/')


def get_back_single_xlsx_file_to_json(xlsx_path: Path, new_json_path: Path) -> Path:
    workbook = load_workbook(filename=xlsx_path)
    sheet = workbook.active
    data = {}

    for row in sheet.iter_rows(values_only=True):
        try:
            data[row[0]] = row[3]
        except Exception as e:
            print(f"Some problem with {xlsx_path}")
            data[row[0]] = row[-1]

    result = unflatten_dict(data)

    with new_json_path.open('w+') as f:
        f.write(json.dumps(result, ensure_ascii=False, indent=4))

    return new_json_path


def get_back_prepared_xlsx_files_to_json(lang_path: Path):
    all_langs_data: [JsonLocalizationDataRouter] = get_json_files_and_path(lang_path)

    for lang_metas in all_langs_data:
        for lang_meta in lang_metas:
            new_json_path = (lang_meta.json_dir_path / f"{lang_meta.json_path.stem}_new.json").resolve()
            get_back_single_xlsx_file_to_json(lang_meta.xlsx_path.resolve(), new_json_path)




# ALL_LANG_PATH = Path("./gmt-landing-locale")

# prepare_default_xlsx_file(ALL_LANG_PATH)
# read_xlsx_file(ALL_LANG_PATH)
# read_xlsx_file_from_folder_and_make_json("gmt-landing-locale")
